import tkinter as tk
from tkinter import messagebox
import sqlite3 as db
import openpyxl
import os
from pathlib import Path
import pandas as pd
import xlsxwriter
from PIL import ImageTk



# create new database
conn = db.connect('pacienti.db')
cur = conn.cursor()
cur.execute(""" CREATE TABLE IF NOT EXISTS pacienti
    (
        rodne_cislo VARCHAR PRIMARY KEY NOT NULL,
        prijmeni VARCHAR NOT NULL,
        jmeno VARCHAR NOT NULL,
        pojistovna INTEGER,
        adresa VARCHAR,
        diagnoza VARCHAR,
        vyska FLOAT,
        vaha FLOAT,
        port BOOLEAN
    ); """)
cur.close()
conn.commit()
conn.close()

# create tkinter window
window = tk.Tk()
window.title('můj malý Štajner')
window.geometry('420x725')
window.config(background='light blue')

# bg = ImageTk.PhotoImage(file = "chemo.png")
bg = ImageTk.PhotoImage(file = "doktor.jpg")
background_label = tk.Label(window, image=bg)
background_label.grid(row=13, column=0, rowspan=1, columnspan=3, pady=5)

title = tk.Label(window, text='Databáza pacientů', font='Arial 16 bold', background='light blue')
title.grid(row=0, column=0, columnspan=3, ipadx=20, ipady=10)

add_frame = tk.LabelFrame(window, text='údaje pacienta', labelanchor='n', background='light blue')
add_frame.grid(row=1, column=0, columnspan=3, padx=5, sticky='ew')


# FUNCTIONS
def add():    # add patient to the database     # nutno poriesit NOT NULL entry!!!
    conn = db.connect('pacienti.db')
    cur = conn.cursor()
    try:
        cur.execute('INSERT INTO pacienti (rodne_cislo, prijmeni, jmeno, pojistovna, adresa, diagnoza, vyska, vaha, port) VALUES(?,?,?,?,?,?,?,?,?)', (fill_id.get(), fill_surname.get(), fill_name.get(), fill_insurance.get(),fill_address.get(), fill_diagnosis.get(), fill_height.get(), fill_weight.get(), fill_port.get()))
        messagebox.showinfo(title='oznameni', message='pacient uspesne pridan do databazy')
    except db.IntegrityError:
        messagebox.showerror(title='Upozornění', message='Pacient s uvedeným rodným číslem již je v databáze!')
    cur.close()
    conn.commit()
    conn.close()
    fill_id.delete(0, 'end')
    fill_surname.delete(0, 'end')
    fill_name.delete(0, 'end')
    fill_insurance.delete(0, 'end')
    fill_address.delete(0, 'end')
    fill_diagnosis.delete(0, 'end')
    fill_height.delete(0, 'end')
    fill_weight.delete(0, 'end')
    fill_port.delete(0, 'end')

def find():    # look for patient in the database, result pops up in a new window
    find_window = tk.Tk()
    find_window.title('pacient')
    find_window.geometry('420x680')
    find_window.config(background='light blue')

    conn = db.connect('pacienti.db')
    cur = conn.cursor()
    id_select = str(fill_search.get())
    cur.execute('SELECT * FROM pacienti WHERE rodne_cislo = ?', ([id_select]))
    record = cur.fetchone()
    note_list = ['rodné číslo', 'příjmení', 'jméno', 'pojišťovna', 'adresa', 'diagnóza', 'výška', 'váha', 'port']
    if record is None:
        messagebox.showinfo(title='Upozornění', message='Pacient s uvedeným rodným číslem není v databáze!')  
        find_window.destroy()
    else:
        for index,item in enumerate(note_list):
            note_label = tk.Label(find_window, width=8, text=item, anchor='w', font='Arial 11', background='light blue')
            note_label.grid(row=index, column=0)
        for i in range(len(note_list)):
            border_label = tk.Label(find_window, width=3, text='>>>', anchor='e', font='Courier 8', background='light blue')
            border_label.grid(row=i, column=1)
        for index, rec in enumerate(record):
            find_label = tk.Label(find_window, width=36, text=rec, anchor='w', font='Arial 11', background='light blue')
            find_label.grid(row=index, column=2)

    conn.commit()
    conn.close()
    return id_select

def delete():    # delete patient from the database
    conn = db.connect('pacienti.db')
    cur = conn.cursor()
    id_select = fill_search.get()
    if messagebox.askyesno(title='Varování', message='Opravdu chceš vymazat pacienta z databázy?') == True:
        cur.execute('DELETE FROM pacienti WHERE rodne_cislo = ?', ([id_select]))
        messagebox.showinfo(title='Info', message='Pacient vymazán!')  
    conn.commit()
    conn.close()
    fill_search.delete(0, 'end')

def save():    # confirm data update and save to the database
    conn = db.connect('pacienti.db')
    cur = conn.cursor()
    id_select = fill_search.get()
    if tk.messagebox.askyesno( title='Varování', message='Opravdu chceš uložit změny?') == True:
        cur.execute('''UPDATE pacienti SET
            rodne_cislo = :rodne_cislo,
            prijmeni = :prijmeni,
            jmeno = :jmeno,
            pojistovna = :pojistovna,
            adresa = :adresa,
            diagnoza = :diagnoza,
            vyska = :vyska,
            vaha = :vaha,
            port = :port
            WHERE rodne_cislo = :rodne_cislo''',
            {
                'prijmeni': fill_surname_edit.get(),
                'jmeno': fill_name_edit.get(),
                'pojistovna' : fill_insurance_edit.get(), 
                'adresa' : fill_address_edit.get(),
                'diagnoza' : fill_diagnosis_edit.get(),
                'vyska' : fill_height_edit.get(),
                'vaha' : fill_weight_edit.get(),
                'port' : fill_port_edit.get(),
                'rodne_cislo' : id_select
            })
        messagebox.showinfo(title='Info', message='Změny uloženy!') 
    conn.commit()
    conn.close()
    edit_window.destroy()

def edit():    # edit patient's data
    global edit_window
    edit_window = tk.Tk()
    edit_window.title('Upravit údaje pacienta')
    edit_window.geometry('420x680')
    edit_window.config(background='light blue')

    edit_title = tk.Label(edit_window, text='Upravit údaje pacienta', font='Arial 16 bold', background='light blue')
    edit_title.grid(row=0, column=0, columnspan=3, ipadx=20, ipady=10)

    global fill_surname_edit
    global fill_name_edit
    global fill_insurance_edit
    global fill_address_edit
    global fill_diagnosis_edit
    global fill_height_edit
    global fill_weight_edit
    global fill_port_edit

    id_edit = tk.Label(edit_window, text='rodné číslo nelze upravit', font='Arial 11', background='light blue')
    id_edit.grid(row=1, column=0)

    surname_edit = tk.Label(edit_window, text='příjmení', font='Arial 11', background='light blue')
    surname_edit.grid(row=2, column=0, pady=2)
    fill_surname_edit = tk.Entry(edit_window, width=20, font='Arial 11', background='light grey')
    fill_surname_edit.grid(row=2, column=1)

    name_edit = tk.Label(edit_window, text='jméno', font='Arial 11', background='light blue')
    name_edit.grid(row=3, column=0, pady=2)
    fill_name_edit = tk.Entry(edit_window, width=20, font='Arial 11', background='light grey')
    fill_name_edit.grid(row=3, column=1)

    insurance_edit = tk.Label(edit_window, text='pojišťovna', font='Arial 11', background='light blue')
    insurance_edit.grid(row=4, column=0, pady=2)
    fill_insurance_edit = tk.Entry(edit_window, width=20, font='Arial 11', background='light grey')
    fill_insurance_edit.grid(row=4, column=1)
 
    address_edit = tk.Label(edit_window, text='adresa', font='Arial 11', background='light blue')
    address_edit.grid(row=5, column=0, pady=2)
    fill_address_edit = tk.Entry(edit_window, width=20, font='Arial 11', background='light grey')
    fill_address_edit.grid(row=5, column=1)

    diagnosis_edit = tk.Label(edit_window, text='diagnóza', font='Arial 11', background='light blue')
    diagnosis_edit.grid(row=6, column=0, pady=2)
    fill_diagnosis_edit = tk.Entry(edit_window, width=20, font='Arial 11', background='light grey')
    fill_diagnosis_edit.grid(row=6, column=1)

    height_edit = tk.Label(edit_window, text='výška', font='Arial 11', background='light blue')
    height_edit.grid(row=7, column=0, pady=2)
    fill_height_edit = tk.Entry(edit_window, width=20, font='Arial 11', background='light grey')
    fill_height_edit.grid(row=7, column=1)

    weight_edit = tk.Label(edit_window, text='váha', font='Arial 11', background='light blue')
    weight_edit.grid(row=8, column=0, pady=2)
    fill_weight_edit = tk.Entry(edit_window, width=20, font='Arial 11', background='light grey')
    fill_weight_edit.grid(row=8, column=1)

    port_edit = tk.Label(edit_window, text='port', font='Arial 11', background='light blue')
    port_edit.grid(row=9, column=0, pady=2)
    fill_port_edit = tk.Entry(edit_window, width=20, font='Arial 11', background='light grey')
    fill_port_edit.grid(row=9, column=1)

    save_btn = tk.Button(edit_window, width=12, text='uložit změny', font='Arial 11', fg='red', command=save)
    save_btn.grid(row=10, column=1, padx=5, pady=10)


    conn = db.connect('pacienti.db')
    cur = conn.cursor()
    id_select = fill_search.get()
    cur.execute('SELECT * FROM pacienti WHERE rodne_cislo = ?', ([id_select]))
    record = cur.fetchone()
    if record is None:
        messagebox.showinfo(title='Upozornění', message='Pacient s uvedeným rodným číslem není v databáze!')
        edit_window.destroy()
    else:
        fill_surname_edit.insert(0, record[1])
        fill_name_edit.insert(0, record[2])
        fill_insurance_edit.insert(0, record[3])
        fill_address_edit.insert(0, record[4])
        fill_diagnosis_edit.insert(0, record[5])
        fill_height_edit.insert(0, record[6])
        fill_weight_edit.insert(0, record[7])
        fill_port_edit.insert(0, record[8])
    conn.commit()
    conn.close()   

def chemo():    # create chemolist (pop up new window)
    chemo_window = tk.Tk()
    chemo_window.title('chemolisty')
    chemo_window.geometry('420x680')
    chemo_window.config(background='light blue')

    var = tk.StringVar(chemo_window)

    def chosen_treatment():    # choose treatment and create notebook
        conn = db.connect('pacienti.db')
        cur = conn.cursor()
        id_select = fill_search.get()
        cur.execute('SELECT * FROM pacienti WHERE rodne_cislo = ?', ([id_select]))
        record = cur.fetchone()
        treatment = var.get()
        try: 
            wb = openpyxl.load_workbook('chemolisty.xlsx')
            ws = wb[treatment]
            ws['A1'].value = record[1]    # update cell values with database data
            ws['A2'].value = record[2]
            ws['A3'].value = record[0]
            ws['A4'].value = record[3]
            ws['A5'].value = record[4]
            ws['A7'].value = record[5]
            ws['C2'].value = record[6]
            ws['D2'].value = record[7]
            if record[8] == 'ano':
                ws['G6'].value = 'PORT'
            if record[8] == 'ne':
                ws['G6'].value = ''
            chemo_file = f'{record[1]}_{record[2]}_{treatment.upper()}.xlsx'

            for sheet in range(len(wb.sheetnames)):    # focus on the working sheet
                if wb.sheetnames[sheet] == treatment:
                    break
            wb.active = sheet

            wb.save(chemo_file)
            messagebox.showinfo(title='Info', message=f'Chemolist úspěšně vytvořen, uložen pod názvem {chemo_file}')
            if messagebox.askyesno(title='Dotaz', message=f'Chceš chemolist otevřít?') == True:
                path_file = Path(chemo_file).resolve()
                os.system(f'start excel.exe "{path_file}"')    # open excel file
            chemo_window.destroy()
        except TypeError:
                messagebox.showinfo(title='Upozornění', message='Pro vytvoření chemolistu musíš zadat rodné číslo!')
                chemo_window.destroy()
        conn.commit()
        conn.close()   

    # dictionary with chemo buttons
    btn_dict = {
        'folfox': 'FOLFOX',
        'fufa': 'FUFA',
        'flot': 'FLOT',
        'carbopt': 'CARBOPLATINA',
        'docetaxel': 'DOCETAXEL'
    }

    for choice, text in btn_dict.items():
        tk.Radiobutton(chemo_window, text=text, variable=var, value=choice, command=chosen_treatment, width=20, background='light yellow', indicator=0).grid(row=list(btn_dict.keys()).index(choice)+1, column=0, padx=10, pady=10, sticky='w')

    tk.Label(chemo_window,text='Vyber požadovaný režim chemoterapie', font='Arial 16 bold', background='light blue').grid(row=0, column=0, columnspan=3, padx=10)


def show_data():    # save all patients from database to excel file
    conn = db.connect('pacienti.db')
    with pd.ExcelWriter("vsichni_pacienti.xlsx", engine="xlsxwriter") as writer:
        try:
            df = pd.read_sql('SELECT * FROM pacienti ORDER BY rodne_cislo ASC', conn)
            df.to_excel(writer, sheet_name = "pacienti", header = True, index = False)
            path_file = Path('vsichni_pacienti.xlsx').resolve()
            os.system(f'start excel.exe "{path_file}"')
        except:
            print("There is an error")


# labels and entries in main window
id = tk.Label(add_frame, text='rodné číslo', font='Arial 11', background='light blue')
id.grid(row=1, column=0)
fill_id = tk.Entry(add_frame, width=20, font='Arial 11', background='light grey')
fill_id.grid(row=1, column=1)
note_id = tk.Label(add_frame, text='bez lomítka', font='Arial 8', background='light blue')
note_id.grid(row=1, column=2)

surname = tk.Label(add_frame, text='příjmení', font='Arial 11', background='light blue')
surname.grid(row=2, column=0, pady=2)
fill_surname = tk.Entry(add_frame, width=20, font='Arial 11', background='light grey')
fill_surname.grid(row=2, column=1)

name = tk.Label(add_frame, text='jméno', font='Arial 11', background='light blue')
name.grid(row=3, column=0, pady=2)
fill_name = tk.Entry(add_frame, width=20, font='Arial 11', background='light grey')
fill_name.grid(row=3, column=1)

insurance = tk.Label(add_frame, text='pojišťovna', font='Arial 11', background='light blue')
insurance.grid(row=4, column=0, pady=2)
fill_insurance = tk.Entry(add_frame, width=20, font='Arial 11', background='light grey')
fill_insurance.grid(row=4, column=1)
note_insurance = tk.Label(add_frame, text='číselný kód', font='Arial 8', background='light blue')
note_insurance.grid(row=4, column=2)

address = tk.Label(add_frame, text='adresa', font='Arial 11', background='light blue')
address.grid(row=5, column=0, pady=2)
fill_address = tk.Entry(add_frame, width=20, font='Arial 11', background='light grey')
fill_address.grid(row=5, column=1)
note_address = tk.Label(add_frame, text='ulice, číslo, PSČ, město', font='Arial 8', background='light blue')
note_address.grid(row=5, column=2)

diagnosis = tk.Label(add_frame, text='diagnóza', font='Arial 11', background='light blue')
diagnosis.grid(row=6, column=0, pady=2)
fill_diagnosis = tk.Entry(add_frame, width=20, font='Arial 11', background='light grey')
fill_diagnosis.grid(row=6, column=1)
note_diagnosis = tk.Label(add_frame, text='MKN-10', font='Arial 8', background='light blue')
note_diagnosis.grid(row=6, column=2)

height = tk.Label(add_frame, text='výška', font='Arial 11', background='light blue')
height.grid(row=7, column=0, pady=2)
fill_height = tk.Entry(add_frame, width=20, font='Arial 11', background='light grey')
fill_height.grid(row=7, column=1)
note_height = tk.Label(add_frame, text='celé číslo v cm', font='Arial 8', background='light blue')
note_height.grid(row=7, column=2)

weight = tk.Label(add_frame, text='váha', font='Arial 11', background='light blue')
weight.grid(row=8, column=0, pady=2)
fill_weight = tk.Entry(add_frame, width=20, font='Arial 11', background='light grey')
fill_weight.grid(row=8, column=1)
note_weight = tk.Label(add_frame, text='celé číslo v kg', font='Arial 8', background='light blue')
note_weight.grid(row=8, column=2)

port = tk.Label(add_frame, text='port', font='Arial 11', background='light blue')
port.grid(row=9, column=0, pady=2)
fill_port = tk.Entry(add_frame, width=20, font='Arial 11', background='light grey')
fill_port.grid(row=9, column=1)
note_port = tk.Label(add_frame, text='ano/ne', font='Arial 8', background='light blue')
note_port.grid(row=9, column=2)

search = tk.Label(window, text='zadej rodné číslo', font='Arial 11', background='light blue')
search.grid(row=11, column=0, pady=20)
fill_search = tk.Entry(window, width=20, font='Arial 11', background='light grey')
fill_search.grid(row=11, column=1)

# BUTTONS
add_btn = tk.Button(add_frame, width=15, text='ulož pacienta', font='Arial 11', fg='green', command=add)
add_btn.grid(row=10, column=1, pady=15)

find_btn = tk.Button(window, width=12, text='zobraz pacienta', font='Arial 11', fg='blue', command=find)
find_btn.grid(row=11, column=2, pady=5)

delete_btn = tk.Button(window, width=12, text='vymaž pacienta', font='Arial 11', fg='red', command=delete)
delete_btn.grid(row=12, column=2, padx=5, pady=5)

edit_btn = tk.Button(window, width=12, text='uprav pacienta', font='Arial 11', fg='brown', command=edit)
edit_btn.grid(row=12, column=0, padx=5, pady=5)

chemo_btn = tk.Button(window, width=12, text='CHEMOLISTY', font='Arial 11', fg='white', background='black', command=chemo)
chemo_btn.grid(row=12, column=1, padx=5, pady=5)

show_data_btn  = tk.Button(window, width=36, text='všichni pacienti - uložit do excelu a zobrazit', font='Arial 11', fg='black', background='wheat', command=show_data)
show_data_btn.grid(row=14, column=0, columnspan=3, padx=5, pady=5)

window.mainloop()